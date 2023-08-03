import openpyxl
from Utility.BaseClass import baseClass


class HomepageData(baseClass):

    @staticmethod
    def getTestData(Test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Abhishek Kulkarni\\Desktop\\AK\\Software Testing\\Automation\\PythonSelFramework\\TestData\\ExcelDemo.xlsx")
        sheet = book.active
        for r in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=r, column=1).value == Test_case_name:
                for c in range(2, sheet.max_column + 1):  # to get column
                    Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
        return[Dict]
        log = baseClass.getLogger()
        log.info(Dict)

