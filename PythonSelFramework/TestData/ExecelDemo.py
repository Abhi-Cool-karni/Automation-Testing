import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Abhishek Kulkarni\\Desktop\\AK\\Software Testing\\Automation\\PythonSelFramework\\TestData\\ExcelDemo.xlsx")
Dict = {}
sheet = book.active
cell = sheet.cell()  # read excel value
# print(cell.value)
# # sheet.cell(row=2, column=2).value = "Abhishek"
# print(sheet.cell(row=2, column=2).value)    # write excel value
# print(sheet.max_row)    # max number of row
# print(sheet.max_column)  # max number of column
# print(sheet['A5'].value)

for r in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=r, column=1).value == "Testcase3":
        for c in range(2, sheet.max_column + 1):  # to get column
            # Dict ["Name"]=" Arvind
            Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
print(Dict)
